# -*- coding: utf-8 -*-
# @Time : 2020/8/12 19:47
# @Author : johnsonLT
# @Site : 
# @File : main.py
# @Software: PyCharm

import sys
import os
import openpyxl
import ctypes
#获取行号列号
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color
#导入QT类
from PyQt5.QtWidgets import QApplication, QMainWindow, QFileDialog, QCalendarWidget
from PyQt5.QtGui import QIcon, QPixmap
#from PyQt5.Qt import QThread, QMutex
from PyQt5.QtCore import pyqtSignal, QObject, QThread
#from PyQt5.QtCore import *
#导入UI界面
import WeeklyReports
from pypinyin import pinyin, Style

class GenerateReportThread(QThread):
    """生成周报线程，继承自QThread"""
    finished_signal = pyqtSignal()
    def __init__(self, directory,path, personReportPathList, curDate):
        super(GenerateReportThread, self).__init__()
        self.directory = directory
        self.summaryReportPath = path
        self.personReportPathList = personReportPathList
        self.curDate = curDate
    def __del__(self):
        self.quit()
        print("calling __del__")
        self.wait()

    def run(self):
        self.on_generateReports()
        self.finished_signal.emit()
        print("GenerateReportThread线程退出")
    '''
    personReportPathList需要传入
    '''
    def on_generateReports(self):
        summaryWorkPath = self.directory + '/' + self.summaryReportPath
        print(summaryWorkPath)
        #self.summaryReportWork = openpyxl.load_workbook(summaryWorkPath)
        self.memberName = []
        for reportWork in self.personReportPathList:
            memberName = self.getStrBetweenSymbol(reportWork, '_', '.')
            self.memberName.append(memberName)
       #按拼音排序
        self.memberName.sort(key=lambda keys:[pinyin(i, style=Style.TONE3) for i in keys])
       #删除sheet页

       #按人名顺序添加sheet页
        for one in self.memberName:
           for reportWork in self.personReportPathList:
               if reportWork.find(one) == -1:
                   continue
               curReportPath = self.directory + '/' + reportWork
               one = self.getStrBetweenSymbol(reportWork, '_', '.')
               #self.listWidget_info.addItem(one)
               self.personReportPathList.remove(reportWork)
               self.del_sheet_by_name(summaryWorkPath, one)
               self.replace_xls(curReportPath, summaryWorkPath, one)
               #break
        #self.listWidget_info.addItem("周报生成完成")
        #生成完成的信号
    '''
    self.personReportWork
    '''
    def del_sheet_by_name(self, tag_file, sheet_name):
        if os.path.exists(tag_file):
            wb = openpyxl.load_workbook(tag_file)
            for name in wb.sheetnames:
                if name.find(sheet_name) == 1:
                    wb.remove(wb[sheet_name])
                    break
            wb.save(filename=tag_file)
        else:
            print("#no target file")
    def getStrBetweenSymbol(self, txt, c_start, c_end):
        start = txt.rfind(c_start)
        if start >= 0:
            start += len(c_start)
            end = txt.find(c_end, start)
            if end >= 0:
                return txt[start:end].strip()
        self.listWidget_info.addItem("日志文件名称错误，姓名前没有下划线")
    '''
    src_file是源xlsx文件，tag_file是目标xlsx文件，sheet_name是目标xlsx里的新sheet名称
    noinspection PyMethodMayBeStatic
    '''
    def replace_xls(self, src_file, tag_file, sheet_name):
        print("Start sheet %s copy from %s to %s" % (sheet_name, src_file, tag_file))
        wbSrc = openpyxl.load_workbook(src_file)
        if os.path.exists(src_file):
            wbTag = openpyxl.load_workbook(tag_file)
        else:
            #创建新表
            wbTag = openpyxl.Workbook()
        if sheet_name in wbSrc.sheetnames:
            wsSrc = wbSrc[sheet_name]
        else:
            wsSrc = wbSrc.active
            wsSrc.title = sheet_name

        if sheet_name in wbTag.sheetnames:
            wsTag = wbTag[sheet_name]
        else:
            wsTag = wbTag.create_sheet(title=sheet_name)
        #处理单个单元格
        self.handleSingleCell(wsSrc, wsTag)
        #处理合并单元格
        self.handleMergedCells(wsSrc, wsTag)
        #冻结单元格
        wsTag.freeze_panes = 'I4'
        wbTag.save(filename=tag_file)
        wbSrc.close()
        wbTag.close()

    '''
    处理合并单元格
    '''
    def handleMergedCells(self, wsSrc, wsTag):
        #print(wsSrc.merged_cells.ranges)  # 获取所有合并单元格
        mergedCellsList = wsSrc.merged_cells.ranges
        maxLen = len(mergedCellsList)
        # wsTag.page_setup.orientation = wsTag.ORIENTATION_LANDSCAPE
        # wsTag.page_setup.paperSize = wsTag.PAPERSIZE_TABLOID
        # wsTag.page_setup.fitToHeight = 3
        # wsTag.page_setup.fitToWidth = 0
        #单元格样式
        #thin = Side(color="000000", border_style="thin")
        medium = Side(color="000000", border_style="medium")
        #thick = Side(color="000000", border_style="thick")
        if len(mergedCellsList) > 0:
            for i in range(0, maxLen):
                #print("mergedCellsList length: %d" % len(mergedCellsList))
                #mergeCells = mergedCellsList[0]
                tagCell = str(mergedCellsList[0])
                cellNum = tagCell.split(":")
                wsSrc.unmerge_cells(tagCell)
                #取行列号
                cell_start_xy = coordinate_from_string(cellNum[0])
                col_num = column_index_from_string(cell_start_xy[0])
                row_start_num = cell_start_xy[1]
                cell_end_xy = coordinate_from_string(cellNum[1])
                row_end_num = cell_end_xy[1]

                try:
                    wsTag.merge_cells(range_string=tagCell)
                    #设置单元格样式
                    #合并单元格的样式和第一个单元格一致
                    focus_cell = wsTag[cellNum[0]]
                    focus_cell.border = Border(top=medium, left=medium, right=medium, bottom=medium)
                    focus_cell.font = Font(name="宋体", color="000000")
                    focus_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    if (row_start_num == 1) or (row_start_num == 2):
                        focus_cell.fill = PatternFill("solid", fgColor="C5D9F1") #蓝色
                    else:
                        focus_cell.fill = PatternFill("solid", fgColor="8DB4E2") #深蓝色
                except (TypeError, ValueError) as e:
                    print(e)
                except:
                    print("Unexpected error:", sys.exc_info()[0])
                    raise
        #处理表头的日期
        wsTag['B1'] = self.curDate

    '''
    处理单个单元格
    '''
    def handleSingleCell(self, wsSrc, wsTag):
        print(wsSrc.rows)
        title_row_num = 3
        thin = Side(color="000000", border_style="thin")
        medium = Side(color="000000", border_style="medium")
        for row in wsSrc.rows:
            for cell in row:
                cell_pos = cell.coordinate
                cell_xy = coordinate_from_string(cell_pos)
                col_num = column_index_from_string(cell_xy[0])
                row_num = cell_xy[1]

                focus_cell = wsTag[cell_pos]
                # 设置单元格样式
                bottom_border_style = Side(color="000000", border_style=cell.border.bottom.border_style)
                left_border_style = Side(color="000000", border_style=cell.border.left.border_style)
                right_border_style = Side(color="000000", border_style=cell.border.right.border_style)
                top_border_style = Side(color="000000", border_style=cell.border.top.border_style)
                focus_cell.border = Border(top=top_border_style, left=left_border_style, right=right_border_style,
                                           bottom=bottom_border_style)
                if cell.value != None:
                    wsTag[cell_pos] = cell.value
                    if cell.value == '周编号':
                        wsTag.merge_cells(range_string = 'A1:A2')
                        focus_cell.fill = PatternFill("solid", fgColor="C5D9F1")  # 蓝色

                    focus_cell.font = Font(name="宋体", color="000000")
                    focus_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    if (row_num == 3) and (cell_pos != 'A3'):
                        focus_cell.fill = PatternFill("solid", fgColor="92D050")  # 绿色
                    elif (row_num > 3) and (col_num > 6):
                        #处理工作内容的边界和颜色
                        focus_cell.fill = PatternFill("solid", fgColor="8DB4E2")  # 深蓝色
                    if (row_num > 3) and (col_num == 4):
                        wsTag[cell_pos].number_format = 'yyyy/mm/dd'
                        wsTag.column_dimensions['D'].width = 12   #这里列号只能写字母，不能写数字
                    if (row_num > 3) and (col_num == 6):
                        wsTag[cell_pos].number_format = '0%'
                else:
                    if (row_num == 3) and (cell_pos != 'A3'):
                        focus_cell.fill = PatternFill("solid", fgColor="92D050")  # 绿色
                        #focus_cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

class AgentUI(QMainWindow, WeeklyReports.Ui_MainWindow):
    def __init__(self):
        super(AgentUI, self).__init__()
        #全局有效的类变量在__init__中声明
        self.curDate : str
        self.initUI()
    def initUI(self):
        self.setupUi(self)  # 向主窗口添加控件
        #设置窗口图标
        icon = QIcon()
        icon.addPixmap(QPixmap('周报面.png'))
        self.setWindowIcon(icon)
        #下面的函数不管参数是什么，任务栏图标都和窗口图标一致，但如果没有下面的函数，任务栏图标就不显示
        ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID()
        #获取当前日期,注意‘月’是大写M
        self.curDate = self.calendarWidget.selectedDate().toString("yyyy年MM月")
        #信号槽连接
        self.btn_select.clicked.connect(self.on_linePathShow)
        self.btn_select.clicked.connect(self.on_filesList)
        #self.btn_generate.clicked.connect(self.on_generateReports)
        self.btn_generate.clicked.connect(self.generateReportThread)

    '''
    slot function, connect to the btn_select clicked signal.
    '''
    def on_linePathShow(self):
        self.directory = QFileDialog.getExistingDirectory(self, "选取文件夹", "./")
        self.line_path.setText(self.directory)
        print(self.directory)
    '''
    slot function, connect to the btn_select clicked signal.
    Description: get the summary report path, and put everyone report path into a list.
    '''
    def on_filesList(self):
        self.filesList = os.listdir(self.directory)
        self.listWidget_filelist.clear()
        self.personReportPathList = []
        for report in self.filesList:
            if report[0] == '.':
                continue
            elif report.find('_') == -1 :
                self.summaryReportPath = report
                print(self.summaryReportPath)
            else:
                self.personReportPathList.append(report)
        self.listWidget_filelist.addItems(self.personReportPathList)

    '''
    slot function, connect to the btn_generate clicked signal.  
    '''
    def generateReportThread(self):
        self.gRthread = GenerateReportThread(self.directory, self.summaryReportPath, self.personReportPathList, self.curDate)
        self.gRthread.finished_signal.connect(lambda:self.slot_showList("周报生成完成"))
        self.gRthread.start()

    def slot_showList(self, data):
        self.listWidget_info.addItem(data)

if __name__ == '__main__':
   app = QApplication(sys.argv)
   main = AgentUI()
   main.show()
   sys.exit(app.exec())
